import tkinter as tk
from tkinter import scrolledtext
import paho.mqtt.client as mqtt_client
import json
from openpyxl import Workbook, load_workbook
import os
import shutil
from datetime import datetime
import threading
import time

def on_connect(client, userdata, flags, rc):
    if rc == 0:
        status_var.set("Connected to MQTT Broker")
    else:
        status_var.set("Failed to connect")

def on_message(client, userdata, msg):
    decoded_payload = msg.payload.decode()
    json_msg = json.loads(decoded_payload)
    payload = json_msg.get("payload", {})
    
    text = None
    if isinstance(payload, str):  # If payload is a string, use it directly
        text = payload
    elif isinstance(payload, dict):  # If payload is a dictionary, get the text key
        text = payload.get("text")

    if text:  # Only append and display the message if the text is found
        received_msg = f"Received: {text}\n"
        msg_history.insert(tk.END, received_msg)
        msg_history.see(tk.END)

        # Append the received message to the Excel file
        append_to_excel("messages_received.xlsx", received_msg)  # Make sure to call 'append_to_excel' here

def connect_mqtt():
    broker = broker_var.get()
    port = int(port_var.get())
    topic = topic_var.get()
    username = username_var.get()
    password = password_var.get()

    client.on_connect = on_connect
    client.on_message = on_message

    client.username_pw_set(username, password)

    # Check if SSL/TLS is enabled and set it up if needed
    if tls_enabled_var.get():
        client.tls_set()

    client.connect(broker, port)
    client.subscribe(topic)

    client.loop_start()

def send_message():
    message = message_entry.get()
    topic = topic_var.get()
    payload = {
        "sender": "broker",
        "type": "sendtext",
        "payload": message
    }
    client.publish(topic, json.dumps(payload))

    message_entry.delete(0, tk.END)

buffered_messages = []

def append_to_excel(filename, content):
    global buffered_messages
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    buffered_messages.append((timestamp, content))  # Include timestamp when buffering messages
    print("Buffered message:", content)
    write_buffered_messages()  # Attempt to write buffered messages

def write_buffered_messages():
    global buffered_messages
    if not buffered_messages:
        return

    try:
        if not os.path.exists("messages_received.xlsx"):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(("Timestamp", "Message"))
            workbook.save("messages_received.xlsx")

        workbook = load_workbook("messages_received.xlsx")
        sheet = workbook.active

        for timestamp, content in buffered_messages:
            sheet.append((timestamp, content))

        workbook.save("messages_received.xlsx")
        buffered_messages = []  # Clear the buffer after successfully writing to the file

    except (PermissionError, OSError) as e:
        print("Error writing to the Excel file:", e)

client = mqtt_client.Client()

root = tk.Tk()
root.title("MQTT Chat")

broker_var = tk.StringVar()
port_var = tk.StringVar()
topic_var = tk.StringVar()
username_var = tk.StringVar()
password_var = tk.StringVar()
status_var = tk.StringVar()

tls_enabled_var = tk.BooleanVar()

topic_var.set("msh/2/json/testnet/!7eff1a90")

tk.Label(root, text="Broker:").grid(row=0, column=0)
tk.Entry(root, textvariable=broker_var).grid(row=0, column=1)

tk.Label(root, text="Port:").grid(row=1, column=0)
tk.Entry(root, textvariable=port_var).grid(row=1, column=1)

tk.Label(root, text="Topic:").grid(row=2, column=0)
tk.Entry(root, textvariable=topic_var).grid(row=2, column=1)

tk.Label(root, text="Username:").grid(row=3, column=0)
tk.Entry(root, textvariable=username_var).grid(row=3, column=1)

tk.Label(root, text="Password:").grid(row=4, column=0)
tk.Entry(root, textvariable=password_var, show="*").grid(row=4, column=1)

tk.Button(root, text="Connect", command=connect_mqtt).grid(row=5, columnspan=2)

tk.Label(root, text="Status:").grid(row=6, column=0)
tk.Label(root, textvariable=status_var).grid(row=6, column=1)

tk.Label(root, text="Message History:").grid(row=7, column=0)
msg_history = scrolledtext.ScrolledText(root, wrap=tk.WORD, width=50, height=10)
msg_history.grid(row=7, column=1)

tk.Label(root, text="Message:").grid(row=8, column=0)
message_entry = tk.Entry(root, width=50)
message_entry.grid(row=8, column=1)

tk.Button(root, text="Send", command=send_message).grid(row=9, columnspan=2)

tk.Checkbutton(root, text="Enable SSL/TLS", variable=tls_enabled_var).grid(row=10, columnspan=2)

root.mainloop()
